import os
import statistics
from datetime import time

import pandas as pd
import openpyxl
import time
from Clases.Validador import Validador
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, Image, PageBreak, Spacer
from PIL import Image as PILImage  # Para obtener las dimensiones originales de la imagen
from reportlab.lib.pagesizes import A4


class Controlador:
    def __init__(self, canti_registros):
        self.ruta_txt = r"C:\Users\Usuario\PycharmProjects\TramasNexus\archivos\MPLUS 8650 QA2.txt"
        self.ruta_txt_result = r"C:\Users\Usuario\PycharmProjects\TramasNexus\archivos\result_8050.txt"
        self.ruta_excel = r"C:\Users\Usuario\PycharmProjects\TramasNexus\archivos\Informacion Trama.xlsx"
        self.ruta_excel_salida = r"C:\Users\Usuario\PycharmProjects\TramasNexus\archivos\Parceador.xlsx"
        self.ruta_pdf_salida = "C:/Users/Usuario/PycharmProjects/TramasNexus/archivos/"
        self.lineas_txt = []
        self.tipo_caracteres = []
        self.requerido = []
        self.largo = []
        self.inicio = []
        self.id_validacion = []
        self.nombre_campo = []
        self.detalleValidacion = []
        self.cant_registros = canti_registros

    def leer_archivo_txt(self):
        try:
            with open(self.ruta_txt, 'r', encoding='utf-8') as archivo:
                self.lineas_txt = archivo.readlines()
            self.lineas_txt = [linea.strip() for linea in self.lineas_txt]
            print("Archivo TXT leído correctamente.")
        except FileNotFoundError:
            print(f"Error: El archivo {self.ruta_txt} no se encontró.")
        except Exception as e:
            print(f"Se produjo un error al leer el archivo TXT: {e}")

    def leer_archivo_excel(self):
        try:
            df = pd.read_excel(self.ruta_excel, engine='openpyxl')

            # Asegúrate de que los índices sean correctos
            self.nombre_campo = df.iloc[:, 1].tolist()  # Columna B (índice 1)
            self.tipo_caracteres = df.iloc[:, 3].tolist()  # Columna G (índice 3)
            self.requerido = df.iloc[:, 4].tolist()  # Columna H (índice 4)
            self.largo = df.iloc[:, 5].tolist()  # Columna I (índice 5)
            self.inicio = df.iloc[:, 6].tolist()  # Columna J (índice 6)
            self.id_validacion = df.iloc[:, 8].tolist()  # Columna I (índice 8)
            self.detalleValidacion = df.iloc[:, 9].tolist()  # Columna J (índice 9)

            print("Archivo Excel leído correctamente.")
        except FileNotFoundError:
            print(f"Error: El archivo {self.ruta_excel} no se encontró.")
        except Exception as e:
            print(f"Se produjo un error al leer el archivo Excel: {e}")

    def obtener_datos_txt(self):
        return self.lineas_txt

    def obtener_datos_excel(self):
        return {
            'tipo_caracteres': self.tipo_caracteres,
            'requerido': self.requerido,
            'largo': self.largo,
            'inicio': self.inicio,
            'id_validacion': self.id_validacion,
            'nombre_campo': self.nombre_campo
        }

    def extraer_subcadenas(self, cadena, inicio, largo):
        """
        Extrae subcadenas de la cadena principal usando las posiciones de inicio y los largos proporcionados.

        :param cadena: La cadena principal de la cual extraer subcadenas.
        :param inicio: Lista de posiciones de inicio para cada subcadena.
        :param largo: Lista de largos de subcadenas a extraer.
        :return: Lista de subcadenas extraídas.
        """
        subcadenas = []

        # Verificar que las listas de inicio y largo tengan la misma longitud
        if len(inicio) != len(largo):
            raise ValueError("Las listas de inicio y largo deben tener la misma longitud.")

        for i in range(len(inicio)):
            start = inicio[i] - 1
            length = largo[i]

            # Asegurarse de que el rango de extracción esté dentro de los límites de la cadena
            if start < 0 or start >= len(cadena):
                subcadenas.append("")  # Agregar una cadena vacía si el índice de inicio está fuera del rango
                continue

            end = min(start + length, len(cadena))  # Ajustar el final si excede la longitud de la cadena
            subcadenas.append(cadena[start:end])

        return subcadenas

    def validar_campos(self):
        inicio = time.time()
        reportes = []
        errore_reportes = []
        list_tipo = self.tipo_caracteres
        list_requerido = self.requerido
        list_id_validacion = self.id_validacion
        list_nombre_campo = self.nombre_campo

        for n in range(self.cant_registros):
            registro1 = self.lineas_txt[n] if self.lineas_txt else ""
            # Dividir el registro en subcadenas
            registro_parseado = self.extraer_subcadenas(registro1, self.inicio, self.largo)
            reporte = []
            for i in range(len(registro_parseado)):
                validador = Validador(registro_parseado[i], list_tipo[i], list_requerido[i], list_id_validacion[i],
                                      list_nombre_campo[i], registro_parseado)
                # Ejecutar todas las validaciones
                es_valido, mensaje = validador.validar_todo(registro_parseado[i])
                if not es_valido:
                    reporte.append(mensaje)
                    # print(mensaje)
            reportes.append(reporte)
            errore_reportes.append(len(reporte))
        media = statistics.mean(errore_reportes)
        # Registrar el tiempo de finalización
        fin = time.time()

        # Calcular el tiempo de ejecución
        tiempo_total = fin - inicio

        print(f"Validación concluida en {tiempo_total:.2f} segundos.:")
        print(f"Total de Registros: {len(self.lineas_txt)}")
        print(f"Registros validados: {self.cant_registros}")
        print(f"Registros con errores: {len(reportes)}")
        print(f"La media de errores por reporte es de: {media}")
        return reportes

    def generar_reporte_pdf(self):
        inicio = time.time()
        # Cargar el archivo Excel
        wb = openpyxl.load_workbook(self.ruta_excel)
        hoja = wb.active

        registros = self.lineas_txt
        validaciones = controlador.validar_campos()

        # Crear un archivo PDF
        nombre_archivo = f"Reporte de Errores de Trama 8650.pdf"
        ruta_completa = os.path.join(self.ruta_pdf_salida, nombre_archivo)
        pdf = SimpleDocTemplate(ruta_completa, pagesize=letter)

        # Crear una lista para almacenar los elementos del PDF
        elementos = []

        # Crear estilos
        styles = getSampleStyleSheet()
        estilo_titulo = styles['Title']
        estilo_titulo.alignment = 1  # Centrar el texto

        # Primera página - Título, imagen y tabla
        # Ruta de la imagen
        image_path = r'C:\Users\Usuario\PycharmProjects\TramasNexus\imagenes\Imagen1.png'

        # Obtener dimensiones originales de la imagen
        with PILImage.open(image_path) as img:
            original_width, original_height = img.size

        # Definir ancho máximo deseado para la imagen en el PDF
        max_width = 4 * inch  # Ajusta el ancho máximo que desees
        # Calcular el alto manteniendo la proporción de la imagen
        aspect_ratio = original_height / original_width
        adjusted_height = max_width * aspect_ratio

        # Insertar la imagen con las dimensiones ajustadas
        img = Image(image_path, width=max_width, height=adjusted_height)
        img.hAlign = 'CENTER'
        elementos.append(Spacer(1, 1 * inch))  # Espacio superior
        elementos.append(img)

        # Espaciado entre imagen y texto
        elementos.append(Spacer(1, 0.5 * inch))

        # Títulos en el documento
        elementos.append(Paragraph("REPORTE DE ERRORES, TRAMA 8650", estilo_titulo))
        elementos.append(Spacer(1, 0.3 * inch))  # Espaciado entre títulos
        elementos.append(Paragraph("PROYECTO: Tramas Nexus - Ciclo 1- Validaciones", estilo_titulo))

        # Espaciado entre título y tabla
        elementos.append(Spacer(1, 0.5 * inch))

        mensaje_estado_ejecucion = "Prueba OK"
        if len(validaciones) > 0:
            mensaje_estado_ejecucion = "Prueba Fallida"

        datos_tabla = [
            ["ID Caso de Prueba", "00001"],
            ["Caso de Prueba", "CP-01 Validación Trama 8650"],
            ["Ciclo de Certificación", "1"],
            ["Estado de Ejecución", mensaje_estado_ejecucion],
            ["Ambiente", "VP-TNR"],
            ["Emisor", "661"],
            ["Ingeniero de pruebas", "Jose Alejandro Perez"]
        ]
        # Crear y estilizar la tabla
        tabla = Table(datos_tabla, colWidths=[200, 200])
        tabla.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 12),
            ('BACKGROUND', (1, 0), (1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        elementos.append(tabla)

        # Espaciado entre tablas
        elementos.append(Spacer(1, 0.5 * inch))

        # Crear la tabla con el Resumen de Validaciones
        datos_tabla2 = [
            ["Trama 8650", ""],
            ["Total de Registros", len(registros)],
            ["Total de Registros Validados", self.cant_registros],
            ["Cantidad de Registros Malos", len(validaciones)],
        ]
        # Crear y estilizar la tabla
        tabla2 = Table(datos_tabla2, colWidths=[200, 200])
        tabla2.setStyle(TableStyle([
            ('SPAN', (0, 0), (-1, 0)),  # Combinar celdas de la primera fila
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),  # Alineación del texto al centro primera fila
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),  # Fondo gris para la cabecera
            ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
            ('TEXTCOLOR', (0, 1), (0, -1), colors.black),
            ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 12),
            ('BACKGROUND', (1, 1), (1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        elementos.append(tabla2)

        # Agregar un salto de página
        elementos.append(PageBreak())

        # Suponiendo que registros es una lista de listas, donde cada lista contiene los errores
        for i, registro in enumerate(validaciones):
            # Espaciado entre tablas
            elementos.append(Spacer(1, 0.5 * inch))

            datos_tabla3 = [
                ["Registro", i + 1],
                ["Total de Campos", 149],
                ["Cantidad de Campos Malos", len(registro)]
            ]
            # Crear y estilizar la tabla
            # Ajustar el ancho de la tabla según el tamaño de la página (en este caso, A4)
            ancho_total = A4[0] - 50  # Resta 50 para dejar márgenes
            ancho_columna1 = ancho_total * 0.50  # 25% del ancho total para la primera columna
            ancho_columna2 = ancho_total * 0.50  # 75% del ancho total para la segunda columna

            tabla3 = Table(datos_tabla3, colWidths=[ancho_columna1, ancho_columna2])
            tabla3.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
                ('TEXTCOLOR', (0, 0), (0, -1), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 12),
                ('BACKGROUND', (1, 0), (1, -1), colors.white),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('WORDWRAP', (0, 0), (-1, -1), True),  # Ajustar el texto dentro de las celdas
            ]))
            elementos.append(tabla3)

            # Crear los datos de la tabla con una cabecera y una fila para cada error
            datos_tabla4 = [
                ["Listado de errores", "", ""],  # Primera fila que abarca todas las columnas
                ["Nombre del campo", "Valor", "Detalle de error"]  # Encabezados de columnas
                  # Otra fila de datos
            ]
            for error in registro:
                datos_tabla4.append([Paragraph(error['nombre_campo']), Paragraph(error['valor']), Paragraph(error['detalle_error'])])

            # Ajustar el ancho de la tabla para que ocupe todo el ancho disponible
            ancho_total = A4[0] - 50  # Resta 50 para dejar márgenes
            ancho_columna1 = ancho_total * 0.30  # 25% del ancho total para la primera columna
            ancho_columna2 = ancho_total * 0.13  # 75% del ancho total para la segunda columna
            ancho_columna3 = ancho_total * 0.57  # 75% del ancho total para la segunda columna

            tabla4 = Table(datos_tabla4, colWidths=[ancho_columna1, ancho_columna2, ancho_columna3])

            # Estilo de la tabla
            tabla4.setStyle(TableStyle([
                ('SPAN', (0, 0), (-1, 0)),  # Combinar celdas de la primera fila
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),  # Alineación del texto al centro primera fila
                ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),  # Fondo gris para la cabecera
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),  # Texto blanco en la cabecera
                ('ALIGN', (0, 1), (-1, -1), 'LEFT'),  # Alineación del texto a la izquierda
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),  # Fuente en negrita para la cabecera
                ('FONTSIZE', (0, 0), (-1, 2), 12),  # Tamaño de fuente cabecera
                ('FONTSIZE', (0, 2), (-1, -1), 8),  # Tamaño de fuente más pequeño
                ('BACKGROUND', (0, 1), (-1, 1), colors.lightgrey),  # Fondo gris claro para los encabezados de columnas
                ('BACKGROUND', (0, 2), (-1, -1), colors.white),  # Fondo beige para el contenido
                ('GRID', (0, 0), (-1, -1), 1, colors.black),  # Bordes de la tabla
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),  # Alineación vertical superior
                ('WORDWRAP', (0, 0), (-1, -1), True),  # Ajuste de texto en celdas
            ]))

            # Añadir la tabla a los elementos del documento
            elementos.append(tabla4)
        # Crear el PDF
        pdf.build(elementos)
        # Registrar el tiempo de finalización
        fin = time.time()

        # Calcular el tiempo de ejecución
        tiempo_total = fin - inicio
        print(f"Archivo Excel '{nombre_archivo}' generado exitosamente en {tiempo_total:.2f} segundos.")



    def generar_parceador(self):

        # Crear un nuevo libro de trabajo
        wb = openpyxl.Workbook()

        # Seleccionar la hoja activa
        hoja = wb.active

        # Escribir los nombres de las columnas en la primera fila
        for col_num, self.nombre_campo in enumerate(self.nombre_campo, start=1):
            hoja.cell(row=1, column=col_num, value=self.nombre_campo)

        datos = controlador.extraer_subcadenas(self.lineas_txt[0], self.inicio, self.largo)
        # Escribir los datos en las filas siguientes
        for col_num, valor in enumerate(datos, start=1):
            hoja.cell(row=2, column=col_num, value=valor)

        # Leer el archivo de texto plano
        with open(self.ruta_txt_result, "r") as archivo:
                lineas = archivo.readlines()

        # Recorrer las líneas y separarlas por punto y coma
        datos_fila = lineas[1].strip().split(';')
        for col_num, valor in enumerate(datos_fila, start=1):
            hoja.cell(row=3, column=col_num, value=valor)



        # Guardar el archivo Excel
        wb.save(self.ruta_excel_salida)

        print(f"Archivo Excel '{self.ruta_excel_salida}' generado exitosamente.")
# Ejemplo de uso:
# Se crea el controlador con 10 registros a validar
controlador = Controlador(10)

# Se leen los documentos de entrada
controlador.leer_archivo_txt()
controlador.leer_archivo_excel()

# Se obtienen los datos leidos
lineas_txt = controlador.obtener_datos_txt()
datos_excel = controlador.obtener_datos_excel()

# Ejemplo de cadena y listas de inicio y largo para parcear un registro
cadena_ejemplo = lineas_txt[0]  # Suponiendo que quieres procesar la primera línea del archivo TXT
inicio = datos_excel['inicio']
largo = datos_excel['largo']
cadena_parceada = controlador.extraer_subcadenas(cadena_ejemplo, inicio, largo)
print(cadena_parceada)


# Generar Reporte
controlador.generar_reporte_pdf()

# Generar Parceador
# controlador.generar_parceador()
